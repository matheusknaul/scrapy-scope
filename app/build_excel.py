from openpyxl import load_workbook
import openpyxl

def __make_excel__(lista):

    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'Produto'
    ws['B1'] = 'Ensaio'
    ws['C1'] = 'Norma'

    for item in range(len(lista)):
        ws[f'A{item + 2}'] = lista[item][0]
        ws[f'B{item + 2}'] = lista[item][1]

    file_name = 'tabela_ensaio.xlsx'

    wb.save(file_name)

    delete_empty(file_name)

    print('Arquivo excel criado com sucesso!')

def delete_empty(excel): # Mudar da coluna B para C
    """Ao deletarmos uma linha no excel usando o openpyxl, ele muda e atualiza instântaneamente
    todos os indices <- precisa-se tomar cuidado com essa característica.

    Args:
        excel (_type_): _description_
    """
    workbook = load_workbook(excel)
    sheet = workbook['Sheet']

    ultima_linha = sheet.max_row
    
    indice_row = ultima_linha
    while indice_row > 0:
        if sheet[f'B{indice_row}'].value == "" or sheet[f'B{indice_row}'].value == " ":
            sheet.delete_rows(indice_row)
        else:
            indice_row -= 1
    workbook.save(excel)